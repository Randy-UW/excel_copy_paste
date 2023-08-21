import openpyxl.utils.datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# -*- coding = utf-8 -*-
# @Time : 2023/7/8 下午 8:34
# @Author : Randy
# @File : Copy_And _Paste.py
# @Software : PyCharm


def copy_and_paste_col(source_file: str, output_file: str,
                       source_sheet='Sheet9', output_sheet='Sheet9',
                       source_range='AA:AC', output_start_col='AD',
                       output_end_col='AF'):
    """
    Depend on openpyxl library
     read and write xlsx file, copy specific column(s) and paste
     to  output_file with specific position
    :param output_end_col: end column in the paste file
    :param output_start_col: start column in the paste file
    :param source_range: range of columns in the copy file
    :param output_sheet: output file's sheet name
    :param source_sheet: source file's sheet name
    :param source_file: source file's name
    :param output_file: output file's name
    :return:
    """
    wb_source = load_workbook(source_file)
    ws_source = wb_source[source_sheet]
    wb_output = load_workbook(output_file)
    ws_output = wb_output[output_sheet]
    # set calendar start from 1904
    wb_source.epoch = \
        openpyxl.utils.datetime.CALENDAR_MAC_1904
    wb_output.epoch = \
        openpyxl.utils.datetime.CALENDAR_MAC_1904
    copy_range = ws_source[source_range]
    row_num = 0
    col_num = column_index_from_string(output_start_col)
    for cell_tuple in copy_range:
        for cell in cell_tuple:
            row_num += 1
            copy_value = cell.value
            ws_output.cell(row=row_num, column=col_num).value \
                = copy_value
        col_num += 1
        row_num = 0
    ws_output['C2']=6
    wb_output.save(output_file)
    wb_source.save(source_file)
